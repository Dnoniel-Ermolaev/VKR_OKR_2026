"""Импорт клинических данных из Excel (multi-sheet) и генерация шаблонов."""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime, date, time
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

from src.infrastructure.db.repository import sql_database_repository
from src.medical.catalog import (
    DIAGNOSIS_BY_ICD,
    LAB_BY_CODE,
    MEDICATION_BY_CODE,
    PROCEDURE_BY_CODE,
    STUDY_BY_CODE,
    VITAL_BY_CODE,
    flag_for_lab,
    flag_for_vital,
    resolve_lab,
    resolve_medication,
    resolve_procedure,
    resolve_study,
    resolve_vital,
)


@dataclass
class SheetSchema:
    name: str
    headers: List[str]
    example: List[Any]
    description: str


SHEETS: Dict[str, SheetSchema] = {
    "Vitals": SheetSchema(
        name="Vitals",
        headers=[
            "recorded_at", "sbp", "dbp", "hr", "rr", "temp", "spo2",
            "diuresis_in", "diuresis_out", "glucose_bedside", "note",
        ],
        example=[
            "2026-03-27T08:00", 135, 85, 88, 18, 36.7, 97, 150, 120, 6.2, "приём",
        ],
        description=(
            "Лист витальных показателей - каждая строка это одно измерение "
            "с временной меткой. Значения вне нормы будут помечены автоматически."
        ),
    ),
    "Labs": SheetSchema(
        name="Labs",
        headers=["recorded_at", "code", "name", "value", "unit", "note"],
        example=["2026-03-27T08:30", "troponin_i", "Тропонин I", 0.02, "нг/мл", ""],
        description=(
            "Лист анализов. Заполните `code` из справочника; если неизвестен - "
            "укажите `name` (например, `Тропонин I`), система попытается найти совпадение."
        ),
    ),
    "Studies": SheetSchema(
        name="Studies",
        headers=["started_at", "completed_at", "code", "name", "status", "priority", "result_text"],
        example=["2026-03-27T09:00", "2026-03-27T09:15", "ecg_12", "ЭКГ 12 каналов", "done", "critical", "синусовый ритм"],
        description="Инструментальные исследования: ЭКГ, ЭхоКГ, СКГ, R-графия и т. д.",
    ),
    "Procedures": SheetSchema(
        name="Procedures",
        headers=["started_at", "completed_at", "code", "name", "status", "operator", "note"],
        example=["2026-03-27T11:00", "2026-03-27T11:40", "pci_stent", "ЧКВ со стентированием", "done", "Иванов И.И.", "2 стента"],
        description="Инвазивные и респираторные процедуры.",
    ),
    "Medications": SheetSchema(
        name="Medications",
        headers=["code", "name", "med_class", "dose", "unit", "route", "frequency", "started_at", "stopped_at", "status"],
        example=["asa", "Ацетилсалициловая кислота", "antiplatelet", "100", "мг", "po", "1 раз/сут", "2026-03-27T08:00", "", "active"],
        description="Активные назначения и курсы препаратов.",
    ),
    "Diagnoses": SheetSchema(
        name="Diagnoses",
        headers=["icd10", "name", "diagnosis_type", "established_at", "note"],
        example=["I21.0", "ОИМ передней стенки", "primary", "2026-03-27T08:00", ""],
        description="Диагнозы по МКБ-10.",
    ),
}


class ExcelImportService:
    """Импорт/экспорт structured данных пациента через Excel (openpyxl)."""

    def __init__(self, db_session: Optional[Session] = None):
        self.db_session = db_session
        self.repo = sql_database_repository(db_session) if db_session is not None else None

    # ------------------------------------------------------------------
    # Templates
    # ------------------------------------------------------------------
    def build_template(self, sheet: str) -> bytes:
        """Сгенерировать xlsx с единственным листом-шаблоном."""
        schema = SHEETS.get(sheet)
        if schema is None:
            raise ValueError(f"Unknown sheet: {sheet}")

        from openpyxl import Workbook  # local import to keep dep optional

        wb = Workbook()
        self._fill_sheet(wb.active, schema)
        # add a README sheet with description
        self._add_readme(wb, [schema])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def build_template_full(self) -> bytes:
        """Сгенерировать полный шаблон со всеми листами."""
        from openpyxl import Workbook

        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for schema in SHEETS.values():
            ws = wb.create_sheet(schema.name)
            self._fill_sheet(ws, schema)
        self._add_readme(wb, list(SHEETS.values()))

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _fill_sheet(self, ws, schema: SheetSchema) -> None:
        ws.title = schema.name
        ws.append(schema.headers)
        if schema.example:
            ws.append(schema.example)
        for idx, header in enumerate(schema.headers, start=1):
            ws.column_dimensions[_col(idx)].width = max(14, min(32, len(str(header)) + 6))

    def _add_readme(self, wb, schemas: List[SheetSchema]) -> None:
        readme = wb.create_sheet("_README", 0)
        readme.append(["Лист", "Колонки", "Описание"])
        for schema in schemas:
            readme.append([
                schema.name,
                ", ".join(schema.headers),
                schema.description,
            ])
        readme.column_dimensions["A"].width = 16
        readme.column_dimensions["B"].width = 60
        readme.column_dimensions["C"].width = 80

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------
    def import_case_data(
        self,
        case_id: str,
        file_bytes: bytes,
        *,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        if self.repo is None:
            raise RuntimeError("ExcelImportService requires a DB session for import")

        case = self.repo.get_case(case_id)
        if case is None:
            return {"error": "Кейс не найден"}

        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

        report: Dict[str, Any] = {
            "imported_total": 0,
            "sheets": {},
            "warnings": [],
            "dry_run": dry_run,
        }

        for sheet_name, schema in SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [self._norm_header(h) for h in rows[0]]
            data_rows = [r for r in rows[1:] if any(cell is not None and str(cell).strip() for cell in r)]
            if not data_rows:
                continue

            imported = 0
            rejected: List[Dict[str, Any]] = []
            for row_idx, row in enumerate(data_rows, start=2):
                record = dict(zip(headers, row))
                try:
                    parsed = self._dispatch(sheet_name, record)
                except ValueError as e:
                    rejected.append({"row": row_idx, "reason": str(e), "raw": _row_preview(record)})
                    continue
                if parsed is None:
                    rejected.append({"row": row_idx, "reason": "пустая строка", "raw": _row_preview(record)})
                    continue
                if not dry_run:
                    try:
                        self._persist(case_id, sheet_name, parsed)
                    except Exception as e:
                        rejected.append({"row": row_idx, "reason": f"db: {e}", "raw": _row_preview(record)})
                        continue
                imported += 1

            report["sheets"][sheet_name] = {
                "imported": imported,
                "rejected": rejected,
                "total_rows": len(data_rows),
            }
            report["imported_total"] += imported

        return report

    # ------------------------------------------------------------------
    # Dispatch per sheet
    # ------------------------------------------------------------------
    def _dispatch(self, sheet: str, record: Dict[str, Any]) -> Optional[Dict[str, Any] | List[Dict[str, Any]]]:
        if sheet == "Vitals":
            return self._parse_vitals_row(record)
        if sheet == "Labs":
            return self._parse_lab_row(record)
        if sheet == "Studies":
            return self._parse_study_row(record)
        if sheet == "Procedures":
            return self._parse_procedure_row(record)
        if sheet == "Medications":
            return self._parse_medication_row(record)
        if sheet == "Diagnoses":
            return self._parse_diagnosis_row(record)
        return None

    def _persist(self, case_id: str, sheet: str, payload: Any) -> None:
        if sheet == "Vitals":
            # payload = list[dict] of observations
            if payload:
                self.repo.add_case_observations(case_id, payload)
            return
        if sheet == "Labs":
            self.repo.add_case_observations(case_id, [payload])
            return
        if sheet == "Studies":
            self.repo.add_case_study(case_id, **payload)
            return
        if sheet == "Procedures":
            self.repo.add_case_procedure(case_id, **payload)
            return
        if sheet == "Medications":
            self.repo.add_case_medication(case_id, **payload)
            return
        if sheet == "Diagnoses":
            self.repo.add_case_diagnosis(case_id, **payload)
            return

    # ------------------------------------------------------------------
    # Row parsers
    # ------------------------------------------------------------------
    def _parse_vitals_row(self, record: Dict[str, Any]) -> List[Dict[str, Any]]:
        recorded_at = self._parse_dt(record.get("recorded_at"))
        observations: List[Dict[str, Any]] = []
        note = _coerce_str(record.get("note"))
        for key in ("sbp", "dbp", "hr", "rr", "temp", "spo2", "diuresis_in", "diuresis_out", "glucose_bedside"):
            if key not in record:
                continue
            value = _coerce_float(record.get(key))
            if value is None:
                continue
            definition = VITAL_BY_CODE.get(key)
            if not definition:
                continue
            observations.append({
                "category": "vital",
                "code": key,
                "name": definition.name_ru,
                "value_num": value,
                "unit": definition.unit,
                "flag": flag_for_vital(key, value),
                "source": "excel",
                "note": note,
                "recorded_at": recorded_at.isoformat() if recorded_at else None,
            })
        if not observations:
            return []
        return observations

    def _parse_lab_row(self, record: Dict[str, Any]) -> Dict[str, Any]:
        recorded_at = self._parse_dt(record.get("recorded_at"))
        code = _coerce_str(record.get("code"))
        name = _coerce_str(record.get("name"))
        resolved = resolve_lab(code) or resolve_lab(name)
        if not resolved:
            raise ValueError(f"не удалось определить код анализа (code={code!r}, name={name!r})")
        definition = LAB_BY_CODE[resolved]
        value_num = _coerce_float(record.get("value"))
        value_text = None if value_num is not None else _coerce_str(record.get("value"))
        unit = _coerce_str(record.get("unit")) or definition.unit
        flag = flag_for_lab(resolved, value_num) if value_num is not None else "unknown"
        return {
            "category": "lab",
            "code": resolved,
            "name": definition.name_ru,
            "value_num": value_num,
            "value_text": value_text,
            "unit": unit,
            "flag": flag,
            "source": "excel",
            "note": _coerce_str(record.get("note")),
            "recorded_at": recorded_at.isoformat() if recorded_at else None,
        }

    def _parse_study_row(self, record: Dict[str, Any]) -> Dict[str, Any]:
        code = _coerce_str(record.get("code"))
        name = _coerce_str(record.get("name"))
        resolved = resolve_study(code) or resolve_study(name)
        if not resolved:
            raise ValueError(f"не удалось определить код исследования (code={code!r}, name={name!r})")
        definition = STUDY_BY_CODE[resolved]
        return {
            "code": resolved,
            "name": definition.name_ru,
            "status": _coerce_str(record.get("status")) or "done",
            "started_at": _iso_or_none(self._parse_dt(record.get("started_at"))),
            "completed_at": _iso_or_none(self._parse_dt(record.get("completed_at"))),
            "result_text": _coerce_str(record.get("result_text")),
            "priority": _coerce_str(record.get("priority")) or definition.acs_priority,
        }

    def _parse_procedure_row(self, record: Dict[str, Any]) -> Dict[str, Any]:
        code = _coerce_str(record.get("code"))
        name = _coerce_str(record.get("name"))
        resolved = resolve_procedure(code) or resolve_procedure(name)
        if not resolved:
            raise ValueError(f"не удалось определить код процедуры (code={code!r}, name={name!r})")
        definition = PROCEDURE_BY_CODE[resolved]
        return {
            "code": resolved,
            "name": definition.name_ru,
            "status": _coerce_str(record.get("status")) or "done",
            "started_at": _iso_or_none(self._parse_dt(record.get("started_at"))),
            "completed_at": _iso_or_none(self._parse_dt(record.get("completed_at"))),
            "operator": _coerce_str(record.get("operator")),
            "priority": _coerce_str(record.get("priority")) or definition.acs_priority,
            "note": _coerce_str(record.get("note")),
        }

    def _parse_medication_row(self, record: Dict[str, Any]) -> Dict[str, Any]:
        code = _coerce_str(record.get("code"))
        name = _coerce_str(record.get("name"))
        resolved = resolve_medication(code) or resolve_medication(name)
        if resolved:
            definition = MEDICATION_BY_CODE[resolved]
            fallback_name = definition.name_ru
            fallback_class = definition.group
            fallback_dose = definition.typical_dose
            fallback_unit = definition.typical_unit
            fallback_route = definition.default_route
        elif name:
            # Free-form medication still accepted, but without catalog link.
            fallback_name = name
            fallback_class = _coerce_str(record.get("med_class"))
            fallback_dose = _coerce_str(record.get("dose"))
            fallback_unit = _coerce_str(record.get("unit"))
            fallback_route = _coerce_str(record.get("route")) or "po"
            resolved = ""
        else:
            raise ValueError(f"не указан препарат (code={code!r}, name={name!r})")

        return {
            "code": resolved,
            "name": name or fallback_name,
            "med_class": _coerce_str(record.get("med_class")) or fallback_class,
            "dose": _coerce_str(record.get("dose")) or fallback_dose,
            "unit": _coerce_str(record.get("unit")) or fallback_unit,
            "route": _coerce_str(record.get("route")) or fallback_route,
            "frequency": _coerce_str(record.get("frequency")),
            "started_at": _iso_or_none(self._parse_dt(record.get("started_at"))),
            "stopped_at": _iso_or_none(self._parse_dt(record.get("stopped_at"))),
            "status": _coerce_str(record.get("status")) or "active",
        }

    def _parse_diagnosis_row(self, record: Dict[str, Any]) -> Dict[str, Any]:
        icd10 = _coerce_str(record.get("icd10")).upper()
        if not icd10:
            raise ValueError("не указан код МКБ-10")
        definition = DIAGNOSIS_BY_ICD.get(icd10)
        name = _coerce_str(record.get("name")) or (definition.name_ru if definition else icd10)
        return {
            "icd10": icd10,
            "name": name,
            "diagnosis_type": _coerce_str(record.get("diagnosis_type")) or "primary",
            "established_at": _iso_or_none(self._parse_dt(record.get("established_at"))),
            "note": _coerce_str(record.get("note")),
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _norm_header(value: Any) -> str:
        return str(value).strip().lower() if value is not None else ""

    @staticmethod
    def _parse_dt(value: Any) -> Optional[datetime]:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        if isinstance(value, (int, float)):
            # Excel serial date - best effort
            try:
                from datetime import timedelta
                base = datetime(1899, 12, 30)
                return base + timedelta(days=float(value))
            except Exception:
                return None
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None


def _col(idx: int) -> str:
    result = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _row_preview(record: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in record.items() if v is not None and str(v).strip()}


def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _iso_or_none(value: Optional[datetime]) -> Optional[str]:
    return value.isoformat() if value else None


__all__ = ["ExcelImportService", "SHEETS"]
