from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.infrastructure.importers.excel_importer import SHEETS


def build_workbook(path: Path, rows_by_sheet: dict[str, list[list[object]]], intro: str) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    readme = wb.create_sheet("_README")
    readme.append(["Лист", "Колонки", "Описание"])
    for schema in SHEETS.values():
        readme.append([schema.name, ", ".join(schema.headers), schema.description])
    readme.append([])
    readme.append(["Комментарий", intro, ""])
    readme.column_dimensions["A"].width = 18
    readme.column_dimensions["B"].width = 90
    readme.column_dimensions["C"].width = 80

    for sheet_name, schema in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(schema.headers)
        for row in rows_by_sheet.get(sheet_name, []):
            ws.append(row)
        for idx, header in enumerate(schema.headers, start=1):
            col = get_column_letter(idx)
            ws.column_dimensions[col].width = max(14, min(32, len(str(header)) + 6))

    try:
        wb.save(path)
        print(f"saved: {path.name}")
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_fixed{path.suffix}")
        wb.save(fallback)
        print(f"saved fallback: {fallback.name}")


def main() -> None:
    structured_rows = {
        "Vitals": [
            ["2026-03-27T08:00", 160, 95, 104, 22, 36.7, 94, 100, 0, 7.1, "первичный осмотр"],
            ["2026-03-27T10:00", 150, 90, 96, 20, 36.6, 95, 200, 50, 6.8, "после начала терапии"],
        ],
        "Labs": [
            ["2026-03-27T08:20", "troponin_i", "Тропонин I", 0.18, "нг/мл", "первое измерение"],
            ["2026-03-27T08:25", "creatinine_blood", "Креатинин", 145, "мкмоль/л", "повышен"],
            ["2026-03-27T08:25", "k_blood", "Калий", 4.8, "ммоль/л", ""],
            ["2026-03-27T08:25", "hgb", "Гемоглобин", 136, "г/л", ""],
        ],
        "Studies": [
            ["2026-03-27T08:30", "2026-03-27T08:40", "ecg_12", "ЭКГ 12-канальная", "done", "critical", "ST-elevation V2-V5"],
            ["2026-03-27T09:30", "2026-03-27T10:10", "echo_cg", "ЭхоКГ", "done", "high", "ФВ 45%, локальная гипокинезия"],
            ["2026-03-27T10:40", "2026-03-27T11:20", "coronary_angiography", "Коронарография (СКГ)", "done", "critical", "окклюзия ПНА"],
        ],
        "Procedures": [
            ["2026-03-27T10:45", "2026-03-27T11:20", "pci_stent", "ЧКВ со стентированием", "done", "Иванов И.И.", "стентирование ПНА"],
            ["2026-03-27T08:05", "2026-03-27T12:00", "oxygen_therapy", "Оксигенотерапия (O2)", "done", "ОРИТ", "5 л/мин"],
        ],
        "Medications": [
            ["asa", "Ацекардол (аспирин)", "antiplatelet", "100", "мг", "po", "1 раз/сут", "2026-03-27T08:10", "", "active"],
            ["ticagrelor", "Brilinta (тикагрелор)", "antiplatelet", "90", "мг", "po", "2 раза/сут", "2026-03-27T08:10", "", "active"],
            ["heparin_uf", "Гепарин (НФГ)", "anticoag", "5000", "ЕД", "iv", "болюс + инфузия", "2026-03-27T08:15", "", "active"],
            ["metoprolol", "Метопролол", "beta_blocker", "50", "мг", "po", "1 раз/сут", "2026-03-27T08:30", "", "active"],
            ["atorvastatin", "Аторвастатин", "statin", "80", "мг", "po", "на ночь", "2026-03-27T08:30", "", "active"],
            ["enalapril", "Эналаприл", "acei", "5", "мг", "po", "2 раза/сут", "2026-03-27T12:00", "", "active"],
        ],
        "Diagnoses": [
            ["I21.0", "Острый трансмуральный инфаркт миокарда передней стенки", "primary", "2026-03-27T08:20", "STEMI"],
            ["I10", "Эссенциальная (первичная) гипертензия", "secondary", "2026-03-27T08:20", ""],
        ],
    }

    protocol_gap_rows = {
        "Vitals": [
            ["2026-03-27T08:00", 155, 92, 98, 20, 36.6, 95, 100, 0, "", "первичный осмотр"],
        ],
        "Labs": [
            ["2026-03-27T08:20", "troponin_i", "Тропонин I", 0.17, "нг/мл", "есть только одно измерение"],
        ],
        "Studies": [
            ["2026-03-27T08:30", "2026-03-27T08:40", "ecg_12", "ЭКГ 12-канальная", "done", "critical", "ST-elevation V2-V5"],
        ],
        "Procedures": [
            ["2026-03-27T08:05", "2026-03-27T12:00", "oxygen_therapy", "Оксигенотерапия (O2)", "done", "ОРИТ", "5 л/мин"],
        ],
        "Medications": [
            ["asa", "Ацекардол (аспирин)", "antiplatelet", "100", "мг", "po", "1 раз/сут", "2026-03-27T08:10", "", "active"],
        ],
        "Diagnoses": [
            ["I21.0", "Острый трансмуральный инфаркт миокарда передней стенки", "primary", "2026-03-27T08:20", "неполный кейс для dashboard"],
        ],
    }

    patient_card_rows = {
        "Vitals": [
            ["2026-01-07T10:55", 135, 92, 78, 18, 36.6, 97, 100, 400, "", "после ЧКВ"],
            ["2026-01-07T18:00", 131, 81, 78, 16, "", 96, "", "", "", "вечер"],
            ["2026-01-08T00:00", 110, 61, 72, 16, "", 96, "", "", "", "ночь"],
            ["2026-01-08T06:00", 129, 80, 70, 16, 36.2, 96, "", 110, "", "утро"],
        ],
        "Labs": [
            ["2026-01-07T11:44", "troponin_i", "Тропонин I", 0.17, "нг/мл", "по карте пациента"],
            ["2026-01-07T15:00", "troponin_i", "Тропонин I", 5.3, "нг/мл", "восстановлено для серийного контроля"],
            ["2026-01-08T08:00", "troponin_i", "Тропонин I", 22.712, "нг/мл", "по переводному эпикризу"],
            ["2026-01-07T11:44", "wbc", "Лейкоциты", 7.42, "10^9/л", "ОАК"],
            ["2026-01-07T11:44", "hgb", "Гемоглобин", 131, "г/л", "ОАК"],
            ["2026-01-07T11:44", "hct", "Гематокрит", 45.6, "%", "ОАК"],
            ["2026-01-07T11:44", "plt", "Тромбоциты", 198, "10^9/л", "ОАК"],
            ["2026-01-07T11:44", "creatinine_blood", "Креатинин", 95, "мкмоль/л", "достроено для протокола"],
            ["2026-01-07T11:44", "k_blood", "Калий", 4.2, "ммоль/л", "достроено для протокола"],
            ["2026-01-07T11:44", "inr", "МНО", 1.0, "", "достроено для протокола"],
            ["2026-01-11T08:00", "cholesterol_total", "Холестерин общий", 3.72, "ммоль/л", "липидный профиль"],
            ["2026-01-11T08:00", "triglycerides", "Триглицериды", 1.11, "ммоль/л", "липидный профиль"],
            ["2026-01-11T08:00", "ldl", "ЛПНП (LDL)", 1.975455, "ммоль/л", "липидный профиль"],
            ["2026-01-11T08:00", "vldl", "ЛПОНП (VLDL)", 0.509174, "ммоль/л", "липидный профиль"],
            ["2026-01-11T08:00", "hdl", "ЛПВП (HDL)", 1.24, "ммоль/л", "липидный профиль"],
            ["2026-01-11T08:00", "atherogenic_index", "Коэффициент атерогенности", 2.0, "", "липидный профиль"],
            ["2026-01-07T12:00", "hbsag", "HbsAg (гепатит B)", "negative", "", "серология"],
            ["2026-01-07T12:00", "rw", "RW (сифилис)", "negative", "", "серология"],
            ["2026-01-07T12:00", "anti_hcv", "Anti-HCV (гепатит C)", "not detected", "", "серология"],
        ],
        "Studies": [
            ["2026-01-07T09:26", "2026-01-07T09:26", "ecg_12", "ЭКГ 12-канальная", "done", "critical", "ST-elevation V2-V5, передний STEMI"],
            ["2026-01-07T11:00", "2026-01-07T11:00", "ecg_12", "ЭКГ 12-канальная", "done", "critical", "синусовый ритм, без существенной отрицательной динамики"],
            ["2026-01-08T06:00", "2026-01-08T06:00", "ecg_12", "ЭКГ 12-канальная", "done", "critical", "синусовый ритм, без существенной динамики"],
            ["2026-01-07T10:20", "2026-01-07T10:55", "coronary_angiography", "Коронарография (СКГ)", "done", "critical", "окклюзия ПНА, выполнена экстренно"],
            ["2026-01-12T09:00", "2026-01-12T09:30", "echo_cg", "ЭхоКГ (ЭхоКС)", "done", "high", "ФВ 48%, локальные нарушения сократимости"],
        ],
        "Procedures": [
            ["2026-01-07T10:20", "2026-01-07T10:55", "pci_stent", "ЧКВ со стентированием", "done", "рентгенэндоваскулярная бригада", "стентирование ПНА"],
            ["2026-01-07T10:20", "2026-01-07T10:55", "mechanical_recanalization", "Механическая реканализация", "done", "рентгенэндоваскулярная бригада", "до стентирования"],
            ["2026-01-07T10:00", "2026-01-07T18:00", "oxygen_therapy", "Оксигенотерапия (O2)", "done", "ОРИТ", "5 л/мин"],
            ["2026-01-07T11:30", "2026-01-08T11:00", "peripheral_catheter", "Постановка ПВК", "done", "медсестра ОРИТ", "по листу наблюдения за катетером"],
        ],
        "Medications": [
            ["asa", "Ацекардол (аспирин)", "antiplatelet", "100", "мг", "po", "1 раз/сут", "2026-01-07T12:00", "", "active"],
            ["ticagrelor", "Brilinta (тикагрелор)", "antiplatelet", "90", "мг", "po", "2 раза/сут", "2026-01-07T12:00", "", "active"],
            ["heparin_uf", "Гепарин (НФГ)", "anticoag", "5000", "ЕД", "iv", "1000 ЕД/час", "2026-01-07T10:20", "", "active"],
            ["metoprolol", "Метопролол", "beta_blocker", "50", "мг", "po", "1 раз/сут", "2026-01-07T12:00", "", "active"],
            ["atorvastatin", "Аторвастатин", "statin", "80", "мг", "po", "на ночь", "2026-01-07T12:00", "", "active"],
            ["enalapril", "Эналаприл", "acei", "5", "мг", "po", "2 раза/сут", "2026-01-07T12:00", "", "active"],
            ["omeprazole", "Омепразол", "ppi", "20", "мг", "po", "1 раз/сут", "2026-01-07T12:00", "", "active"],
        ],
        "Diagnoses": [
            ["I21.0", "Острый трансмуральный инфаркт миокарда передней стенки", "primary", "2026-01-07T10:05", "по карте пациента"],
            ["I10", "Эссенциальная (первичная) гипертензия", "secondary", "2026-01-07T10:05", "сопутствующий диагноз"],
            ["I50.9", "ХСН неуточненная", "secondary", "2026-01-07T10:05", "сопутствующий диагноз"],
        ],
    }

    build_workbook(
        BASE_DIR / "structured_high_risk_demo.xlsx",
        structured_rows,
        "Короткий импортируемый кейс высокого риска для презентации.",
    )
    build_workbook(
        BASE_DIR / "protocol_gap_demo.xlsx",
        protocol_gap_rows,
        "Неполный STEMI-кейс для демонстрации protocol-driven dashboard и pending шагов.",
    )
    build_workbook(
        BASE_DIR / "patient_card_stemi_case.xlsx",
        patient_card_rows,
        "Кейс, собранный по реальной карте пациента из Итог.png и аккуратно достроенный для полного structured-case сценария.",
    )

    print("Demo Excel files regenerated.")


if __name__ == "__main__":
    main()
